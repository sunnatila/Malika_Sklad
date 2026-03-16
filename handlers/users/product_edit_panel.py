import os
import tempfile

from aiogram.types import Message, CallbackQuery, FSInputFile
from aiogram.fsm.context import FSMContext
from aiogram.exceptions import TelegramBadRequest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .start import AdminFilter
from loader import dp, db
from states import ProductReduce, ProductIncrease, BrandChange
from keyboards.default import product_buttons, product_type_buttons
from keyboards.inline import (
    product_list_keyboard, product_detail_keyboard,
    product_delete_confirm_keyboard, action_cancel_keyboard,
    brand_filter_keyboard, brand_change_keyboard
)

# ==================== KONFIGURATSIYA ====================

TYPE_CONFIG = {
    "batareyka": {
        "prefix": "bat", "name": "Batareyka", "plural": "Batareykalar", "emoji": "🔋",
        "table": "batteries", "has_brand": True,
        "headers": ['#', 'Nomi', 'Brand', 'Soni', "Sana"],
        "widths": [5, 28, 16, 8, 14],
        "soni_col": 4,
        "row_fn": lambda p, i: [
            i, p['title'], p.get('category', '') or '',
            p['count'], p['created_at'].strftime('%d.%m.%Y') if p.get('created_at') else ''
        ],
        "get_all": "get_all_batteries", "get_by_id": "get_battery_by_id",
        "get_by_cat": "get_batteries_by_category", "get_cats": "get_battery_categories",
    },
    "zaryadka": {
        "prefix": "chr", "name": "Zaryadka", "plural": "Zaryadkalar", "emoji": "🔌",
        "table": "chargers", "has_brand": True,
        "headers": ['#', 'Nomi', 'Brand', 'Quvvat', 'Kuchlanish', 'Soni', "Sana"],
        "widths": [5, 28, 16, 10, 12, 8, 14],
        "soni_col": 6,
        "row_fn": lambda p, i: [
            i, p['title'], p.get('category', '') or '',
            p.get('watt', '') or '', p.get('voltage', '') or '',
            p['count'], p['created_at'].strftime('%d.%m.%Y') if p.get('created_at') else ''
        ],
        "get_all": "get_all_chargers", "get_by_id": "get_charger_by_id",
        "get_by_cat": "get_chargers_by_category", "get_cats": "get_charger_categories",
    },
    "display": {
        "prefix": "dsp", "name": "Display", "plural": "Displaylar", "emoji": "🖥",
        "table": "displays", "has_brand": False,
        "headers": ['#', 'Nomi', 'Hz', 'Pin', 'Soni', "Sana"],
        "widths": [5, 28, 10, 10, 8, 14],
        "soni_col": 5,
        "row_fn": lambda p, i: [
            i, p['title'], p.get('hz', '') or '', p.get('pin', '') or '',
            p['count'], p['created_at'].strftime('%d.%m.%Y') if p.get('created_at') else ''
        ],
        "get_all": "get_all_displays", "get_by_id": "get_display_by_id",
    },
}

PREFIX_MAP = {v["prefix"]: k for k, v in TYPE_CONFIG.items()}


def get_config_by_prefix(prefix):
    ptype = PREFIX_MAP.get(prefix)
    return TYPE_CONFIG[ptype] if ptype else None


# ==================== YORDAMCHI ====================

def format_detail(product, ptype) -> str:
    cfg = TYPE_CONFIG[ptype]
    text = (
        f"📋 <b>Mahsulot haqida to'liq ma'lumot:</b>\n\n"
        f"🆔 ID: <b>{product['id']}</b>\n"
        f"{cfg['emoji']} Tur: <b>{cfg['name']}</b>\n"
        f"📝 Nomi: <b>{product['title']}</b>\n"
    )
    if product.get('category'):
        text += f"🏷 Brand: <b>{product['category']}</b>\n"
    if product.get('voltage'):
        text += f"🔌 Kuchlanish: <b>{product['voltage']}</b>\n"
    if product.get('watt'):
        text += f"⚡ Quvvat: <b>{product['watt']}</b>\n"
    if product.get('hz'):
        text += f"📺 Chastota: <b>{product['hz']}</b>\n"
    if product.get('pin'):
        text += f"🔌 Pin: <b>{product['pin']}</b>\n"
    text += f"📦 Soni: <b>{product['count']}</b> dona\n"
    if product.get('created_at'):
        text += f"📅 Qo'shilgan: <b>{product['created_at'].strftime('%d.%m.%Y')}</b>\n"
    return text


def generate_excel(products, cfg, brand_name=None) -> str:
    wb = Workbook()
    ws = wb.active
    sheet_title = f"{cfg['plural']}"
    if brand_name:
        sheet_title += f" ({brand_name})"
    ws.title = sheet_title[:31]  # Excel sheet nomi 31 ta belgigacha

    hfont = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    hfill = PatternFill('solid', fgColor='2E86AB')
    dfont = Font(name='Arial', size=10)
    brd = Border(
        left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
    )
    ca = Alignment(horizontal='center', vertical='center')
    la = Alignment(horizontal='left', vertical='center')

    for col, h in enumerate(cfg['headers'], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hfont, hfill, ca, brd

    for idx, p in enumerate(products, 1):
        row = idx + 1
        row_data = cfg['row_fn'](p, idx)
        rf = PatternFill('solid', fgColor='F0F8FF' if idx % 2 == 0 else 'FFFFFF')
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font, c.border, c.fill = dfont, brd, rf
            c.alignment = ca if col in [1, cfg['soni_col']] else la

    tr = len(products) + 2
    tf = PatternFill('solid', fgColor='E8F5E9')
    tbf = Font(name='Arial', bold=True, size=10)
    for col in range(1, len(cfg['headers']) + 1):
        c = ws.cell(row=tr, column=col, value='')
        c.fill, c.border = tf, brd
    ws.cell(row=tr, column=2, value='JAMI:').font = tbf
    sl = ws.cell(row=1, column=cfg['soni_col']).column_letter
    tc = ws.cell(row=tr, column=cfg['soni_col'])
    tc.value, tc.font, tc.alignment = f'=SUM({sl}2:{sl}{tr-1})', tbf, ca

    for i, w in enumerate(cfg['widths'], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = 'A2'

    prefix = brand_name or cfg['plural']
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', prefix=f'{prefix}_', delete=False, dir=tempfile.gettempdir())
    fp = tmp.name
    tmp.close()
    wb.save(fp)
    return fp


# ==================== YORDAMCHI: mahsulotlar + excel yuborish ====================

async def send_product_list_with_excel(msg_or_call, products, cfg, prefix, brand_name=None):
    """Mahsulotlar ro'yxati va Excel ni yuborish (msg yoki call uchun)"""
    is_call = isinstance(msg_or_call, CallbackQuery)

    title_text = f"{cfg['emoji']} <b>{cfg['plural']}</b>"
    if brand_name:
        title_text += f" — 🏷 <b>{brand_name}</b>"
    title_text += f" ({len(products)} ta):\n\nMahsulotni tanlang 👇"

    kb = product_list_keyboard(products, prefix, brand_name)

    if is_call:
        try:
            await msg_or_call.message.edit_text(title_text, reply_markup=kb)
        except TelegramBadRequest:
            pass  # xabar o'zgarmagan — e'tibor bermaymiz
        msg_target = msg_or_call.message
    else:
        await msg_or_call.answer(title_text, reply_markup=kb)
        msg_target = msg_or_call

    # Excel
    try:
        fp = generate_excel(products, cfg, brand_name)
        fname = f"{cfg['plural']}"
        if brand_name:
            fname += f"_{brand_name}"
        doc = FSInputFile(fp, filename=f"{fname}.xlsx")
        caption = f"📊 {cfg['emoji']} <b>{cfg['plural']}</b>"
        if brand_name:
            caption += f" ({brand_name})"
        caption += f" — {len(products)} ta mahsulot"
        await msg_target.answer_document(doc, caption=caption)
        os.unlink(fp)
    except Exception as e:
        await msg_target.answer(f"❗ Excel xatolik: {str(e)}")


# ======================== RO'YXAT ========================

@dp.message(AdminFilter(), lambda msg: msg.text == "📋 Mahsulotlar ro'yxati")
async def product_list_start(msg: Message, state: FSMContext):
    await msg.answer("Qaysi turdagi mahsulotlarni ko'rmoqchisiz?", reply_markup=product_type_buttons())
    await state.set_state("product_list_type")


@dp.message(AdminFilter(), lambda msg: msg.text in ["🔋 Batareyka", "🔌 Zaryadka", "🖥 Display"])
async def product_list_by_type(msg: Message, state: FSMContext):
    current = await state.get_state()
    if current != "product_list_type":
        return

    tmap = {"🔋 Batareyka": "batareyka", "🔌 Zaryadka": "zaryadka", "🖥 Display": "display"}
    ptype = tmap[msg.text]
    cfg = TYPE_CONFIG[ptype]
    prefix = cfg['prefix']

    # Display uchun brand yo'q — to'g'ridan-to'g'ri ko'rsatish
    if not cfg.get('has_brand'):
        products = await getattr(db, cfg['get_all'])()
        if not products:
            await state.clear()
            return await msg.answer(f"❗ {cfg['emoji']} {cfg['plural']} topilmadi", reply_markup=product_buttons())

        await send_product_list_with_excel(msg, products, cfg, prefix)
        await state.clear()
        await msg.answer("⬇️", reply_markup=product_buttons())
        return

    # Batareyka/Zaryadka — avval brand tanlash
    categories = await getattr(db, cfg['get_cats'])()

    if not categories:
        # Brandlar yo'q — to'g'ridan-to'g'ri barcha mahsulotlarni ko'rsatish
        products = await getattr(db, cfg['get_all'])()
        if not products:
            await state.clear()
            return await msg.answer(f"❗ {cfg['emoji']} {cfg['plural']} topilmadi", reply_markup=product_buttons())

        await send_product_list_with_excel(msg, products, cfg, prefix)
        await state.clear()
        await msg.answer("⬇️", reply_markup=product_buttons())
        return

    # Brandlarni ko'rsatish
    total = await getattr(db, cfg['get_all'])()
    await msg.answer(
        f"{cfg['emoji']} <b>{cfg['plural']}</b> ({len(total)} ta)\n\n🏷 Brand bo'yicha tanlang yoki barchasini ko'ring:",
        reply_markup=brand_filter_keyboard(categories, prefix)
    )
    await state.clear()
    await msg.answer("⬇️", reply_markup=product_buttons())


# ======================== BRAND FILTER — inline callback ========================

@dp.callback_query(lambda c: c.data.split("_")[1] == "brand")
async def brand_filter_chosen(call: CallbackQuery):
    """Brand tanlanganda — shu brand bo'yicha mahsulotlar"""
    parts = call.data.split("_")
    prefix = parts[0]
    brand = "_".join(parts[2:])  # brand nomida _ bo'lishi mumkin
    cfg = get_config_by_prefix(prefix)
    if not cfg:
        return

    if brand == "all":
        # Barchasi
        products = await getattr(db, cfg['get_all'])()
        if not products:
            return await call.answer("❗ Mahsulotlar topilmadi", show_alert=True)
        await send_product_list_with_excel(call, products, cfg, prefix)
    else:
        # Brandga qarab
        products = await getattr(db, cfg['get_by_cat'])(brand)
        if not products:
            return await call.answer(f"❗ {brand} brandi bo'yicha mahsulot topilmadi", show_alert=True)
        await send_product_list_with_excel(call, products, cfg, prefix, brand)


# ======================== BRAND FILTER — ortga qaytish ========================

@dp.callback_query(lambda c: c.data.split("_")[1] == "brandback")
async def brand_back(call: CallbackQuery):
    """Mahsulotlar ro'yxatidan brand tanlashga qaytish"""
    prefix = call.data.split("_")[0]
    cfg = get_config_by_prefix(prefix)
    if not cfg:
        return

    categories = await getattr(db, cfg['get_cats'])()
    total = await getattr(db, cfg['get_all'])()

    if not categories:
        # Brandlar yo'q — barchasini ko'rsatish
        if not total:
            return await call.answer("❗ Mahsulotlar topilmadi", show_alert=True)
        await send_product_list_with_excel(call, total, cfg, prefix)
        return

    try:
        await call.message.edit_text(
            f"{cfg['emoji']} <b>{cfg['plural']}</b> ({len(total)} ta)\n\n🏷 Brand bo'yicha tanlang yoki barchasini ko'ring:",
            reply_markup=brand_filter_keyboard(categories, prefix)
        )
    except TelegramBadRequest:
        await call.answer("🏷 Siz allaqachon brandlar sahifasidasiz", show_alert=False)


# ======================== TAFSILOT ========================

@dp.callback_query(lambda c: c.data.split("_")[1] == "view")
async def product_view(call: CallbackQuery):
    parts = call.data.split("_")
    prefix, pid = parts[0], int(parts[2])
    cfg = get_config_by_prefix(prefix)
    if not cfg:
        return
    product = await getattr(db, cfg['get_by_id'])(pid)
    if not product:
        return await call.answer("❗ Topilmadi", show_alert=True)
    await call.message.edit_text(
        format_detail(product, PREFIX_MAP[prefix]),
        reply_markup=product_detail_keyboard(pid, prefix)
    )


# ======================== ORTGA (mahsulot detaildan) ========================

@dp.callback_query(lambda c: c.data.split("_")[1] == "back")
async def product_back(call: CallbackQuery):
    prefix = call.data.split("_")[0]
    cfg = get_config_by_prefix(prefix)
    if not cfg:
        return

    # Brand bor turlar uchun — brand tanlashga qaytish
    if cfg.get('has_brand'):
        categories = await getattr(db, cfg['get_cats'])()
        total = await getattr(db, cfg['get_all'])()

        if categories:
            try:
                await call.message.edit_text(
                    f"{cfg['emoji']} <b>{cfg['plural']}</b> ({len(total)} ta)\n\n🏷 Brand bo'yicha tanlang yoki barchasini ko'ring:",
                    reply_markup=brand_filter_keyboard(categories, prefix)
                )
            except TelegramBadRequest:
                pass
            return

    # Brand yo'q yoki display — to'g'ridan-to'g'ri ro'yxat
    products = await getattr(db, cfg['get_all'])()
    try:
        await call.message.edit_text(
            f"{cfg['emoji']} <b>{cfg['plural']}</b> ({len(products)} ta):\n\nMahsulotni tanlang 👇",
            reply_markup=product_list_keyboard(products, prefix)
        )
    except TelegramBadRequest:
        pass


# ======================== BRAND O'ZGARTIRISH ========================

@dp.callback_query(lambda c: c.data.split("_")[1] == "brchg")
async def brand_change_start(call: CallbackQuery):
    """Brand o'zgartirish boshlash — inline brand tugmalarini ko'rsatish"""
    parts = call.data.split("_")
    prefix, pid = parts[0], int(parts[2])
    cfg = get_config_by_prefix(prefix)
    if not cfg:
        return
    product = await getattr(db, cfg['get_by_id'])(pid)
    if not product:
        return await call.answer("❗ Topilmadi", show_alert=True)

    current_brand = product.get('category', '') or "belgilanmagan"
    await call.message.edit_text(
        f"🏷 <b>Brand o'zgartirish</b>\n\n"
        f"📝 Mahsulot: <b>{product['title']}</b>\n"
        f"🏷 Hozirgi brand: <b>{current_brand}</b>\n\n"
        f"Yangi brandni tanlang:",
        reply_markup=brand_change_keyboard(pid, prefix)
    )


@dp.callback_query(lambda c: c.data.split("_")[1] == "brnew")
async def brand_change_select(call: CallbackQuery):
    """Brand tanlash — DB da yangilash"""
    parts = call.data.split("_")
    prefix, pid = parts[0], int(parts[2])
    new_brand = parts[3] if len(parts) > 3 else ""
    cfg = get_config_by_prefix(prefix)
    if not cfg:
        return

    if new_brand == "none":
        new_brand = ""

    try:
        await db.update_category(cfg['table'], pid, new_brand)
        product = await getattr(db, cfg['get_by_id'])(pid)
        if not product:
            return await call.answer("❗ Topilmadi", show_alert=True)

        brand_text = new_brand if new_brand else "o'chirildi"
        await call.answer(f"✅ Brand: {brand_text}", show_alert=True)
        await call.message.edit_text(
            format_detail(product, PREFIX_MAP[prefix]),
            reply_markup=product_detail_keyboard(pid, prefix)
        )
    except Exception as e:
        await call.answer(f"❌ Xatolik: {str(e)}", show_alert=True)


@dp.callback_query(lambda c: c.data.split("_")[1] == "brcancel")
async def brand_change_cancel(call: CallbackQuery):
    """Brand o'zgartirish bekor qilish — mahsulot tafsilotiga qaytish"""
    parts = call.data.split("_")
    prefix, pid = parts[0], int(parts[2])
    cfg = get_config_by_prefix(prefix)
    if not cfg:
        return
    product = await getattr(db, cfg['get_by_id'])(pid)
    if not product:
        return await call.answer("❗ Topilmadi", show_alert=True)
    await call.message.edit_text(
        format_detail(product, PREFIX_MAP[prefix]),
        reply_markup=product_detail_keyboard(pid, prefix)
    )


# ======================== QO'SHISH (stock) ========================

@dp.callback_query(lambda c: c.data.split("_")[1] == "inccancel")
async def inc_cancel(call: CallbackQuery, state: FSMContext):
    parts = call.data.split("_")
    prefix, pid = parts[0], int(parts[2])
    cfg = get_config_by_prefix(prefix)
    await state.clear()
    product = await getattr(db, cfg['get_by_id'])(pid)
    if not product:
        return await call.answer("❗ Topilmadi", show_alert=True)
    await call.message.edit_text(
        format_detail(product, PREFIX_MAP[prefix]),
        reply_markup=product_detail_keyboard(pid, prefix)
    )


@dp.callback_query(lambda c: c.data.split("_")[1] == "inc")
async def inc_start(call: CallbackQuery, state: FSMContext):
    parts = call.data.split("_")
    prefix, pid = parts[0], int(parts[2])
    cfg = get_config_by_prefix(prefix)
    product = await getattr(db, cfg['get_by_id'])(pid)
    if not product:
        return await call.answer("❗ Topilmadi", show_alert=True)
    await state.set_state(ProductIncrease.amount)
    await state.update_data(inc_id=pid, inc_prefix=prefix, inc_table=cfg['table'])
    await call.message.edit_text(
        f"📥 <b>Qo'shish</b>\n\n📝 {product['title']}\n📦 Hozirgi: <b>{product['count']}</b> dona\n\nNechta qo'shilganini kiriting:",
        reply_markup=action_cancel_keyboard(pid, prefix, "inc")
    )


@dp.message(ProductIncrease.amount)
async def inc_amount(msg: Message, state: FSMContext):
    if not msg.text.isdigit():
        return await msg.answer("❗ Faqat son kiriting!")
    amount = int(msg.text)
    if amount <= 0:
        return await msg.answer("❗ Son 0 dan katta bo'lishi kerak!")
    data = await state.get_data()
    pid, prefix, table = data['inc_id'], data['inc_prefix'], data['inc_table']
    cfg = get_config_by_prefix(prefix)
    try:
        new_count = await db.increase_count(table, pid, amount)
        await state.clear()
        await msg.answer(f"✅ Qo'shildi!\n\n📥 Qo'shildi: <b>{amount}</b>\n📦 Jami: <b>{new_count}</b> dona")
        product = await getattr(db, cfg['get_by_id'])(pid)
        if product:
            await msg.answer(
                format_detail(product, PREFIX_MAP[prefix]),
                reply_markup=product_detail_keyboard(pid, prefix)
            )
    except Exception as e:
        await state.clear()
        await msg.answer(f"❌ Xatolik: {str(e)}")
        await msg.answer("📦 Mahsulotlar bo'limi", reply_markup=product_buttons())


# ======================== CHIQARISH ========================

@dp.callback_query(lambda c: c.data.split("_")[1] == "deccancel")
async def dec_cancel(call: CallbackQuery, state: FSMContext):
    parts = call.data.split("_")
    prefix, pid = parts[0], int(parts[2])
    cfg = get_config_by_prefix(prefix)
    await state.clear()
    product = await getattr(db, cfg['get_by_id'])(pid)
    if not product:
        return await call.answer("❗ Topilmadi", show_alert=True)
    await call.message.edit_text(
        format_detail(product, PREFIX_MAP[prefix]),
        reply_markup=product_detail_keyboard(pid, prefix)
    )


@dp.callback_query(lambda c: c.data.split("_")[1] == "dec")
async def dec_start(call: CallbackQuery, state: FSMContext):
    parts = call.data.split("_")
    prefix, pid = parts[0], int(parts[2])
    cfg = get_config_by_prefix(prefix)
    product = await getattr(db, cfg['get_by_id'])(pid)
    if not product:
        return await call.answer("❗ Topilmadi", show_alert=True)
    if product['count'] == 0:
        return await call.answer("❗ Soni 0, chiqarish mumkin emas!", show_alert=True)
    await state.set_state(ProductReduce.amount)
    await state.update_data(dec_id=pid, dec_prefix=prefix, dec_table=cfg['table'], dec_count=product['count'])
    await call.message.edit_text(
        f"📤 <b>Chiqarish</b>\n\n📝 {product['title']}\n📦 Hozirgi: <b>{product['count']}</b> dona\n\n"
        f"Nechta chiqarilganini kiriting (1 dan {product['count']} gacha):",
        reply_markup=action_cancel_keyboard(pid, prefix, "dec")
    )


@dp.message(ProductReduce.amount)
async def dec_amount(msg: Message, state: FSMContext):
    if not msg.text.isdigit():
        return await msg.answer("❗ Faqat son kiriting!")
    amount = int(msg.text)
    if amount <= 0:
        return await msg.answer("❗ Son 0 dan katta bo'lishi kerak!")
    data = await state.get_data()
    pid, prefix, table, pcount = data['dec_id'], data['dec_prefix'], data['dec_table'], data['dec_count']
    cfg = get_config_by_prefix(prefix)
    if amount > pcount:
        return await msg.answer(f"❗ Omborda faqat <b>{pcount}</b> ta bor. 1 dan {pcount} gacha kiriting.")
    try:
        new_count = await db.reduce_count(table, pid, amount)
        await state.clear()
        await msg.answer(f"✅ Chiqarildi!\n\n📤 Chiqarildi: <b>{amount}</b>\n📦 Qoldi: <b>{new_count}</b> dona")
        product = await getattr(db, cfg['get_by_id'])(pid)
        if product:
            await msg.answer(
                format_detail(product, PREFIX_MAP[prefix]),
                reply_markup=product_detail_keyboard(pid, prefix)
            )
    except Exception as e:
        await state.clear()
        await msg.answer(f"❌ Xatolik: {str(e)}")
        await msg.answer("📦 Mahsulotlar bo'limi", reply_markup=product_buttons())


# ======================== O'CHIRISH ========================

@dp.callback_query(lambda c: c.data.split("_")[1] == "delyes")
async def del_yes(call: CallbackQuery):
    parts = call.data.split("_")
    prefix, pid = parts[0], int(parts[2])
    cfg = get_config_by_prefix(prefix)
    try:
        await db.delete_product(cfg['table'], pid)
        await call.message.edit_text("✅ Mahsulot o'chirildi!")
        await call.message.answer("📦 Mahsulotlar bo'limi", reply_markup=product_buttons())
    except Exception as e:
        await call.message.edit_text(f"❌ Xatolik: {str(e)}")


@dp.callback_query(lambda c: c.data.split("_")[1] == "delno")
async def del_no(call: CallbackQuery):
    parts = call.data.split("_")
    prefix, pid = parts[0], int(parts[2])
    cfg = get_config_by_prefix(prefix)
    product = await getattr(db, cfg['get_by_id'])(pid)
    if not product:
        return await call.answer("❗ Topilmadi", show_alert=True)
    await call.message.edit_text(
        format_detail(product, PREFIX_MAP[prefix]),
        reply_markup=product_detail_keyboard(pid, prefix)
    )


@dp.callback_query(lambda c: c.data.split("_")[1] == "del")
async def del_start(call: CallbackQuery):
    parts = call.data.split("_")
    prefix, pid = parts[0], int(parts[2])
    cfg = get_config_by_prefix(prefix)
    product = await getattr(db, cfg['get_by_id'])(pid)
    if not product:
        return await call.answer("❗ Topilmadi", show_alert=True)
    await call.message.edit_text(
        f"🗑 <b>O'chirib tashlash</b>\n\n📝 <b>{product['title']}</b> ni o'chirmoqchimisiz?\n\n⚠️ Bu amalni qaytarib bo'lmaydi!",
        reply_markup=product_delete_confirm_keyboard(pid, prefix)
    )